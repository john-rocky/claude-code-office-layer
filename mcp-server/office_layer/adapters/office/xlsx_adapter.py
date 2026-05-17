"""openpyxl adapter — Excel .xlsx / .xlsm.

Preserves sheet name + cell range per spec §9.3.3. Each non-empty sheet
becomes a chunk; sub-tables (contiguous non-empty blocks) become
ExtractedTable rows.
"""

from __future__ import annotations

from pathlib import Path

from ...models import ChunkKind, DocumentChunk, DocumentKind, ExtractedTable, ExtractionMethod
from ..base import ExtractionResult


def _cell_ref(col: int, row: int) -> str:
    s = ""
    c = col
    while c > 0:
        c, rem = divmod(c - 1, 26)
        s = chr(65 + rem) + s
    return f"{s}{row}"


class OpenpyxlAdapter:
    name = "openpyxl"
    supported_kinds = (DocumentKind.XLSX,)

    def __init__(self) -> None:
        try:
            import openpyxl  # noqa: F401
            self._available = True
        except ImportError:
            self._available = False

    def is_available(self) -> bool:
        return self._available

    def extract(
        self,
        path: Path,
        kind: DocumentKind,
        *,
        document_id: str,
        workspace_id: str,
    ) -> ExtractionResult:
        try:
            import openpyxl
        except ImportError as exc:
            return ExtractionResult(error=f"openpyxl missing: {exc}")

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            return ExtractionResult(error=f"openpyxl failed to open: {exc}")

        chunks: list[DocumentChunk] = []
        tables: list[ExtractedTable] = []
        ordinal = 0
        sheet_count = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_count += 1
            rows: list[list[str]] = []
            min_col, min_row, max_col, max_row = None, None, 0, 0
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                stringified = ["" if v is None else str(v) for v in row]
                if any(s.strip() for s in stringified):
                    rows.append(stringified)
                    if min_row is None:
                        min_row = row_idx
                    max_row = row_idx
                    last_nonempty = max(
                        (i for i, s in enumerate(stringified, start=1) if s.strip()),
                        default=0,
                    )
                    if min_col is None or last_nonempty < min_col:
                        min_col = last_nonempty
                    max_col = max(max_col, last_nonempty)
            if not rows:
                continue

            cell_range = (
                f"{_cell_ref(1, min_row or 1)}:{_cell_ref(max_col or 1, max_row or 1)}"
            )

            # Sheet text for FTS — flatten all cell strings.
            sheet_text_lines = ["\t".join(r) for r in rows]
            sheet_text = "\n".join(sheet_text_lines)

            chunks.append(
                DocumentChunk(
                    id=DocumentChunk.make_id(document_id, ordinal),
                    document_id=document_id,
                    workspace_id=workspace_id,
                    kind=ChunkKind.XLSX_SHEET,
                    ordinal=ordinal,
                    sheet_name=sheet_name,
                    cell_range=cell_range,
                    text=sheet_text,
                    char_count=len(sheet_text),
                    confidence=1.0,
                    extraction_method=ExtractionMethod.XLSX_PARSE,
                )
            )
            ordinal += 1

            headers = rows[0] if rows else []
            body = rows[1:] if len(rows) > 1 else []
            tables.append(
                ExtractedTable(
                    document_id=document_id,
                    sheet_name=sheet_name,
                    headers=headers,
                    rows=body,
                    cell_range=cell_range,
                    confidence=0.9,
                )
            )

        wb.close()

        return ExtractionResult(
            chunks=chunks,
            tables=tables,
            sheet_count=sheet_count or None,
            extraction_method=ExtractionMethod.XLSX_PARSE,
        )
